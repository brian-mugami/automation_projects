import os
import pdfkit
from PyPDF2 import PdfMerger
from xlsx2html import xlsx2html
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook


def optimize_column_widths(sheet):
    """Optimize column widths based on the max length of the data, ignoring merged cells."""
    for column in sheet.columns:
        max_length = 0
        column_letter = None

        # Find the first non-merged cell in the column to get the column letter
        for cell in column:
            if not any(cell.coordinate in merged_range for merged_range in sheet.merged_cells.ranges):
                column_letter = cell.column_letter  # Get the column letter
                break  # Exit loop after finding the first non-merged cell

        # Calculate max length of non-merged cells in the column
        for cell in column:
            if not any(cell.coordinate in merged_range for merged_range in sheet.merged_cells.ranges):
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

        # Adjust the column width based on the maximum length found
        if column_letter is not None:
            adjusted_width = (max_length + 2)  # Add some padding
            sheet.column_dimensions[column_letter].width = adjusted_width


def is_sheet_content_present(sheet):
    """Check if the sheet contains any non-empty cells."""
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):  # Check if there's at least one non-empty cell
            return True
    return False


def convert_excel_to_pdf(excel_path, output_pdf, path='wkhtmltopdf.exe'):
    workbook = load_workbook(excel_path)
    merger = PdfMerger()

    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path)

    # Iterate through each sheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Check if there is content in the sheet
        if not is_sheet_content_present(sheet):
            print(f"Skipping empty sheet: {sheet_name}")
            continue  # Skip empty sheets

        # Optimize column widths
        optimize_column_widths(sheet)

        # Save the optimized workbook to a temporary file
        temp_workbook_path = f"optimized_{sheet_name}.xlsx"
        workbook.save(temp_workbook_path)

        # Create a temporary HTML file to preserve formatting
        with NamedTemporaryFile(suffix='.html', delete=False) as html_file:
            # Convert the optimized Excel to HTML
            xlsx2html(temp_workbook_path, html_file.name, sheet=sheet_name)

            # Set PDF options for landscape orientation and margins
            pdf_options = {
                'orientation': 'Landscape',
                'page-size': 'A4',
                'margin-top': '10mm',
                'margin-right': '10mm',
                'margin-bottom': '10mm',
                'margin-left': '10mm',
                'no-stop-slow-scripts': '',
                'disable-smart-shrinking': ''
            }

            # Convert the HTML to a PDF using pdfkit with landscape orientation
            with NamedTemporaryFile(suffix='.pdf', delete=False) as pdf_file:
                pdfkit.from_file(html_file.name, pdf_file.name, options=pdf_options, configuration=config)
                merger.append(pdf_file.name)

        # Remove the temporary optimized workbook
        os.remove(temp_workbook_path)

    # Write the final merged PDF
    with open(output_pdf, 'wb') as final_output:
        merger.write(final_output)
        merger.close()

    print(f"Converted Excel workbook to PDF: {output_pdf}")
