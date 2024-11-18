import os

import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

NO_TABLES = "No tables in the given range"


class TableException(Exception):
    def __init__(self, message: str):
        super().__init__(message)


def extract_tables_from_pdf(pdf_path, from_page=None, to_page=None):
    tables = []
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        from_page = from_page if from_page is not None else 1
        to_page = to_page if to_page is not None else total_pages

        for i in range(from_page - 1, to_page):
            page = pdf.pages[i]
            page_tables = page.extract_tables()
            for table in page_tables:
                if table:
                    tables.append((i + 1, table))
    if len(tables) == 0:
        raise TableException(NO_TABLES)
    return tables


def create_excel_with_tables(tables, excel_path):
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)
    for page_number, table in tables:
        df = pd.DataFrame(table[1:], columns=table[0])
        sheet_name = f"Page_{page_number}_Table"
        while sheet_name in load_workbook(excel_path).sheetnames:
            sheet_name += "_1"
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


