import os
from concurrent.futures import ThreadPoolExecutor
from tempfile import NamedTemporaryFile

import pdfkit
from PyPDF2 import PdfMerger
from openpyxl import load_workbook
from xlsx2html import xlsx2html


def optimize_column_widths(sheet):
    """Optimize column widths based on the max length of the data, ignoring merged cells."""
    for column in sheet.columns:
        max_length = 0
        column_letter = None

        for cell in column:
            if not any(cell.coordinate in merged_range for merged_range in sheet.merged_cells.ranges):
                column_letter = cell.column_letter
                break

        for cell in column:
            if not any(cell.coordinate in merged_range for merged_range in sheet.merged_cells.ranges):
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

        if column_letter is not None:
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width


def is_sheet_content_present(sheet):
    """Check if the sheet contains any non-empty cells."""
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            return True
    return False


def process_sheet_to_pdf(sheet_name, workbook, config, merger):
    """Process a single sheet into PDF."""
    sheet = workbook[sheet_name]

    if not is_sheet_content_present(sheet):
        print(f"Skipping empty sheet: {sheet_name}")
        return

    optimize_column_widths(sheet)

    temp_workbook_path = f"optimized_{sheet_name}.xlsx"
    workbook.save(temp_workbook_path)

    with NamedTemporaryFile(suffix='.html', delete=False) as html_file:
        xlsx2html(temp_workbook_path, html_file.name, sheet=sheet_name)

        pdf_options = {
            'orientation': 'Landscape',
            'page-size': 'A4',
            'margin-top': '10mm',
            'margin-right': '10mm',
            'margin-bottom': '10mm',
            'margin-left': '10mm',
        }

        with NamedTemporaryFile(suffix='.pdf', delete=False) as pdf_file:
            pdfkit.from_file(html_file.name, pdf_file.name, options=pdf_options, configuration=config)
            merger.append(pdf_file.name)

    os.remove(temp_workbook_path)


def convert_excel_to_pdf(excel_path, output_pdf_path, num_threads=4,
                         path='C:/Users/CHC18/Desktop/new_ideas (1)/projects/wkhtmltopdf.exe'):
    workbook = load_workbook(excel_path)
    merger = PdfMerger()

    path_wkhtmltopdf = path
    config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)

    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        futures = [executor.submit(process_sheet_to_pdf, sheet_name, workbook, config, merger) for sheet_name in
                   workbook.sheetnames]

        # Wait for all threads to complete
        for future in futures:
            future.result()

    with open(output_pdf_path, 'wb') as final_output:
        merger.write(final_output)
        merger.close()

    print(f"Converted Excel workbook to PDF: {output_pdf_path}")

# Example usage
# input = "C:/Users/CHC18/Desktop/new_ideas (1)/projects/FUNCTIONAL REQUIREMENTS_TREASURY_ v1.2.1.xlsx"
# convert_excel_to_pdf(input, 'output_document.pdf', num_threads=4)
